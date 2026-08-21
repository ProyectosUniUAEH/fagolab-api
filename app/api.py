"""Rutas HTTP de fago-api. Todas bajo el prefijo /api."""
from __future__ import annotations

import hashlib
import io
import os
import re
import uuid
from datetime import date

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse, Response

from . import repo, repo_mol, admin, repo_biblioteca
from .config import settings

router = APIRouter(prefix="/api")


def _guard_data_ops(request: Request) -> None:
    """En local, cualquiera autenticado con permiso. En Kaanbal, solo superadmin."""
    if settings.ENV.lower() in {"local", "dev", "development"}:
        return
    user = getattr(request.state, "user", None) or {}
    if user.get("isSuperadmin"):
        return
    raise HTTPException(
        403,
        "En el laboratorio solo la superadministradora puede sembrar, limpiar o restaurar datos.",
    )


# ---- Catálogos / proyecto -------------------------------------------------------------
@router.get("/proyecto")
def proyecto():
    return repo.get_proyecto()


@router.get("/dashboard")
def dashboard():
    return repo.dashboard()


@router.get("/catalogos/organos")
def catalogo_organos():
    return repo.list_organos()


@router.post("/catalogos/organos")
def crear_organo(payload: dict):
    try:
        return repo.crear_organo(payload)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/catalogos/medios")
def catalogo_medios():
    return repo.list_medios()


@router.get("/catalogos/preferencias-siembra")
def preferencias_siembra():
    return repo.list_preferencias_siembra()


@router.post("/catalogos/medios")
def crear_medio(payload: dict):
    try:
        return repo.crear_medio(payload)
    except ValueError as e:
        raise HTTPException(400, str(e))


# ---- Lecturas de la cadena ------------------------------------------------------------
@router.get("/recepciones")
def recepciones():
    return repo.list_recepciones()


@router.get("/peces")
def peces():
    return repo.list_peces()


@router.get("/muestras")
def muestras():
    return repo.list_muestras()


@router.get("/cajas")
def cajas():
    return repo.list_cajas()


@router.get("/subcultivos")
def subcultivos():
    return repo.list_subcultivos()


@router.patch("/subcultivos/{id_sub}")
def actualizar_subcultivo(id_sub: str, payload: dict):
    try:
        return repo.update_subcultivo(id_sub, payload)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.patch("/cajas/{id_caja}/trazabilidad")
def actualizar_caja_trazabilidad(id_caja: str, payload: dict):
    try:
        return repo.update_caja_trazabilidad(id_caja, payload.get("trazabilidad"))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/extracciones")
def extracciones():
    return repo.list_extracciones()


@router.get("/viales")
def viales():
    return repo.list_viales()


@router.get("/nanodrop")
def nanodrop():
    return repo.list_nanodrop()


@router.patch("/nanodrop/{id_lectura}")
def actualizar_nanodrop(id_lectura: str, payload: dict):
    try:
        return repo.update_nanodrop(id_lectura, payload)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/pcr")
def pcr():
    return repo.list_pcr()


@router.get("/pcr/pendientes")
def pcr_pendientes():
    """Cola de viales aprobados en NanoDrop para preparar una corrida de PCR."""
    return repo.list_viales_pendientes_pcr()


@router.get("/geles")
def geles():
    return repo.list_geles()


@router.get("/etiquetas")
def etiquetas():
    return repo.list_etiquetas()


# ---- Reportes (exportación tipo Excel de Pamela) --------------------------------------
@router.get("/reportes/resumen")
def reporte_resumen():
    """Datos JSON con la forma de la hoja Resumen del Excel."""
    return repo.reporte_rows()


@router.get("/reportes/excel")
def reporte_excel():
    """Descarga un .xlsx con la forma de la hoja 'Resumen' del Excel de muestreo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = repo.reporte_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    headers = ["Lote", "Nº (F)", "Muestra", "Organo", "Medio de cultivo",
               "Descripción", "260/280", "260/230", "ng/uL", "Muestreo"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1F4E5F")
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    last_lote = None
    for r in rows:
        lote = r.get("lote") or ""
        # como en el Excel: el Lote solo se escribe cuando cambia
        ws.append([
            lote if lote != last_lote else "",
            r.get("frasco") or "",
            r.get("muestra") or "",
            r.get("organo") or "",
            r.get("medio") or "",
            r.get("descripcion") or "",
            r.get("r280"), r.get("r230"), r.get("ngul"),
            r.get("muestreo") or "",
        ])
        last_lote = lote
    widths = [8, 10, 12, 12, 16, 28, 9, 9, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Muestreos_peces_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _texto_pdf(valor: object) -> str:
    """Texto seguro para el PDF sencillo de reportes (fuente Helvetica base)."""
    texto = str(valor if valor not in (None, "") else "-")
    return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)").replace("\n", " ")


def _crear_pdf_reporte(rows: list[dict]) -> bytes:
    """Genera un PDF portable sin depender de datos de ejemplo ni librerias extra."""
    total = len(rows)
    con_nanodrop = sum(1 for row in rows if row.get("ngul") is not None)
    aptas = sum(1 for row in rows if str(row.get("muestreo") or "").lower().find("pcr") >= 0)
    fecha = date.today().isoformat()
    paginas: list[list[str]] = [[
        "FagoLab | Reporte de trazabilidad experimental",
        f"Generado: {fecha}",
        "",
        f"Registros consolidados: {total}",
        f"Lecturas NanoDrop disponibles: {con_nanodrop}",
        f"Registros con decisión hacia PCR: {aptas}",
        "",
        "Este documento resume los registros capturados en el sistema.",
        "Las lecturas NanoDrop orientan la decisión; no garantizan el resultado de PCR.",
        "",
        "Detalle de muestras",
        "Lote | Frasco | Muestra | Órgano | Medio | 260/280 | 260/230 | ng/uL | Decisión",
    ]]
    for indice in range(0, total, 30):
        if indice:
            paginas.append([
                "FagoLab | Detalle de muestras (continuación)",
                "Lote | Frasco | Muestra | Órgano | Medio | 260/280 | 260/230 | ng/uL | Decisión",
            ])
        for row in rows[indice:indice + 30]:
            paginas[-1].append(" | ".join([
                _texto_pdf(row.get("lote")), _texto_pdf(row.get("frasco")), _texto_pdf(row.get("muestra")),
                _texto_pdf(row.get("organo")), _texto_pdf(row.get("medio")), _texto_pdf(row.get("r280")),
                _texto_pdf(row.get("r230")), _texto_pdf(row.get("ngul")), _texto_pdf(row.get("muestreo")),
            ]))

    objetos: list[bytes] = [b"", b"", b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"]
    ids_paginas: list[int] = []
    for lineas in paginas:
        contenido = ["BT", "/F1 10 Tf", "50 760 Td", "14 TL"]
        for posicion, linea in enumerate(lineas):
            if posicion in (0, 9):
                contenido.append("/F1 13 Tf" if posicion == 0 else "/F1 10 Tf")
            elif posicion == 1:
                contenido.append("/F1 9 Tf")
            contenido.append(f"({_texto_pdf(linea)}) Tj")
            contenido.append("T*")
        stream = "\n".join(contenido + ["ET"]).encode("latin-1", "replace")
        contenido_id = len(objetos) + 1
        objetos.append(f"<< /Length {len(stream)} >>\nstream\n".encode() + stream + b"\nendstream")
        pagina_id = len(objetos) + 1
        objetos.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 3 0 R >> >> /Contents {contenido_id} 0 R >>".encode())
        ids_paginas.append(pagina_id)

    objetos[0] = b"<< /Type /Catalog /Pages 2 0 R >>"
    hijos = " ".join(f"{identificador} 0 R" for identificador in ids_paginas)
    objetos[1] = f"<< /Type /Pages /Kids [{hijos}] /Count {len(ids_paginas)} >>".encode()
    salida = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for numero, objeto in enumerate(objetos, start=1):
        offsets.append(len(salida))
        salida.extend(f"{numero} 0 obj\n".encode())
        salida.extend(objeto)
        salida.extend(b"\nendobj\n")
    inicio_xref = len(salida)
    salida.extend(f"xref\n0 {len(objetos) + 1}\n0000000000 65535 f \n".encode())
    for offset in offsets[1:]: salida.extend(f"{offset:010d} 00000 n \n".encode())
    salida.extend(f"trailer\n<< /Size {len(objetos) + 1} /Root 1 0 R >>\nstartxref\n{inicio_xref}\n%%EOF".encode())
    return bytes(salida)


@router.get("/reportes/pdf")
def reporte_pdf():
    """Descarga un reporte PDF con resumen y registros reales del laboratorio."""
    fname = f"Reporte_FagoLab_{date.today().isoformat()}.pdf"
    return Response(
        _crear_pdf_reporte(repo.reporte_rows()),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---- Escrituras -----------------------------------------------------------------------
@router.post("/recepciones")
def crear_recepcion(payload: dict):
    return repo.crear_recepcion(payload)


@router.patch("/recepciones/{id_recepcion}")
def actualizar_recepcion(id_recepcion: str, payload: dict):
    try:
        return repo.actualizar_recepcion(id_recepcion, payload)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.post("/aislamiento")
def registrar_aislamiento(payload: dict):
    if not payload.get("idRecepcion"):
        raise HTTPException(400, "idRecepcion es obligatorio")
    if not payload.get("organos") or not payload.get("medios"):
        raise HTTPException(400, "Selecciona al menos un órgano y un medio")
    return repo.registrar_aislamiento(payload)


@router.post("/cajas/{id_caja}/observacion")
def observacion_caja(id_caja: str, payload: dict):
    return repo.registrar_observacion(id_caja, payload)


@router.post("/cajas/{id_caja}/subcultivo")
def subcultivo_de_caja(id_caja: str, payload: dict):
    return repo_mol.crear_subcultivo(id_caja, payload)


@router.post("/cajas/{id_caja}/subcultivos")
def subcultivos_de_caja(id_caja: str, payload: dict):
    """Registra varias colonias de la misma caja de una sola vez."""
    try:
        return repo_mol.crear_subcultivos(id_caja, payload)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/subcultivos/{id_sub}/extraccion")
def extraccion_de_subcultivo(id_sub: str, payload: dict):
    return repo_mol.crear_extraccion(id_sub, payload)


@router.post("/viales/{id_vial}/nanodrop")
def nanodrop_de_vial(id_vial: str, payload: dict):
    return repo_mol.crear_nanodrop(id_vial, payload)


@router.post("/viales/{id_vial}/pcr")
def pcr_de_vial(id_vial: str, payload: dict):
    return repo_mol.crear_pcr(id_vial, payload)


@router.post("/geles")
def crear_gel(payload: dict):
    return repo_mol.crear_gel(payload)


# ---- Caja de controles positivos ------------------------------------------------------
@router.get("/positivos")
def positivos():
    return repo_mol.list_controles_positivos()


@router.post("/positivos")
def crear_positivo(payload: dict):
    if not payload.get("etiqueta"):
        raise HTTPException(400, "La etiqueta del control positivo es obligatoria")
    return repo_mol.crear_control_positivo(payload)


@router.patch("/positivos/{id_control}")
def actualizar_positivo(id_control: str, payload: dict):
    return repo_mol.actualizar_control_positivo(id_control, payload)


# ---- Corridas de PCR (memoria de muestras / lotes amplificados) -----------------------
@router.get("/pcr/corridas")
def corridas_pcr():
    return repo_mol.list_corridas_pcr()


@router.post("/pcr/corridas")
def crear_corrida_pcr(payload: dict):
    if not payload.get("muestras"):
        raise HTTPException(400, "Selecciona al menos una muestra para la corrida")
    return repo_mol.crear_corrida_pcr(payload)


@router.get("/pcr/corridas/{id_corrida}/pozos")
def pozos_sugeridos(id_corrida: str):
    """Plantilla de pozos para un gel (blanco, control +, muestras) desde una corrida."""
    try:
        return repo_mol.generar_pozos_desde_corrida(id_corrida)
    except ValueError as e:
        raise HTTPException(404, str(e))


# ---- Media (imágenes) -----------------------------------------------------------------
_SAFE = re.compile(r"[^A-Za-z0-9._-]")


@router.post("/media")
async def subir_media(
    file: UploadFile = File(...),
    codigoObjeto: str = Form(...),
    rol: str = Form("evidencia"),
    esPrincipal: bool = Form(False),
):
    os.makedirs(settings.MEDIA_DIR, exist_ok=True)
    raw = await file.read()
    sha = hashlib.sha256(raw).hexdigest()
    ext = os.path.splitext(file.filename or "")[1].lower() or ".bin"
    safe = _SAFE.sub("_", os.path.splitext(file.filename or "img")[0])[:40]
    fname = f"{uuid.uuid4().hex[:12]}_{safe}{ext}"
    path = os.path.join(settings.MEDIA_DIR, fname)
    with open(path, "wb") as f:
        f.write(raw)
    storage_uri = f"/media/{fname}"
    try:
        res = repo_mol.guardar_media_vinculo(
            storage_uri, file.filename or fname, file.content_type or "image/*",
            len(raw), sha, codigoObjeto, rol, bool(esPrincipal),
        )
    except ValueError as e:
        raise HTTPException(404, str(e))
    return res


@router.get("/media/objeto/{codigo}")
def media_de_objeto(codigo: str):
    return repo_mol.list_media_de_objeto(codigo)


# ---- Biblioteca (bibliografía en PDF) -------------------------------------------------
@router.get("/biblioteca")
def biblioteca(categoria: str | None = None):
    return repo_biblioteca.listar(categoria)


@router.post("/biblioteca")
async def crear_documento_biblioteca(
    titulo: str = Form(...),
    autores: str = Form(""),
    anio: str = Form(""),
    fuente: str = Form(""),
    categoria: str = Form("referencia_doctora"),
    notas: str = Form(""),
    doi: str = Form(""),
    file: UploadFile | None = File(None),
):
    archivo_uri = archivo_nombre = None
    size_bytes = None
    if file is not None and file.filename:
        os.makedirs(repo_biblioteca.BIBLIO_DIR, exist_ok=True)
        raw = await file.read()
        ext = os.path.splitext(file.filename)[1].lower() or ".pdf"
        if ext != ".pdf":
            raise HTTPException(400, "Solo se aceptan archivos PDF.")
        safe = _SAFE.sub("_", os.path.splitext(file.filename)[0])[:60]
        fname = f"{uuid.uuid4().hex[:12]}_{safe}{ext}"
        with open(os.path.join(repo_biblioteca.BIBLIO_DIR, fname), "wb") as f:
            f.write(raw)
        archivo_uri = f"/biblioteca/{fname}"
        archivo_nombre = file.filename
        size_bytes = len(raw)
    payload = {"titulo": titulo, "autores": autores, "anio": anio, "fuente": fuente,
               "categoria": categoria, "notas": notas, "doi": doi}
    try:
        return repo_biblioteca.crear(payload, archivo_uri, archivo_nombre, size_bytes)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.patch("/biblioteca/{id_documento}")
def actualizar_documento_biblioteca(id_documento: str, payload: dict):
    try:
        return repo_biblioteca.actualizar(id_documento, payload)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.delete("/biblioteca/{id_documento}")
def borrar_documento_biblioteca(id_documento: str):
    try:
        return repo_biblioteca.eliminar(id_documento)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.delete("/media/{id_media}")
def borrar_media(id_media: str):
    try:
        return repo.eliminar_media(id_media)
    except ValueError as e:
        raise HTTPException(404, str(e))


# ---- Borrado seguro de registros ------------------------------------------------------
@router.delete("/cajas/{id_caja}")
def borrar_caja(id_caja: str):
    try:
        return repo.eliminar_caja(id_caja)
    except ValueError as e:
        raise HTTPException(409, str(e))


@router.delete("/subcultivos/{id_sub}")
def borrar_subcultivo(id_sub: str):
    try:
        return repo.eliminar_subcultivo(id_sub)
    except ValueError as e:
        raise HTTPException(409, str(e))


# ---- Control de datos: pruebas y backups ---------------------------------------------
@router.get("/admin/estado")
def admin_estado():
    payload = admin.estado()
    payload["entorno"] = settings.ENV
    payload["opsHabilitadas"] = settings.ENV.lower() in {"local", "dev", "development"}
    return payload


@router.get("/admin/exportar")
def admin_exportar():
    """Descarga un .sql con todos los datos actuales (backup auto-restaurable)."""
    sql = admin.exportar_sql()
    stamp = date.today().isoformat()
    return Response(
        content=sql,
        media_type="application/sql; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="fagolab-backup-{stamp}.sql"'},
    )


@router.post("/admin/limpiar")
def admin_limpiar(request: Request):
    _guard_data_ops(request)
    return admin.limpiar()


@router.post("/admin/sembrar")
def admin_sembrar(request: Request):
    _guard_data_ops(request)
    try:
        return admin.sembrar()
    except FileNotFoundError as e:
        raise HTTPException(409, str(e))


@router.post("/admin/guardar-semilla")
def admin_guardar_semilla(request: Request):
    _guard_data_ops(request)
    return admin.guardar_semilla()


@router.post("/admin/importar")
async def admin_importar(request: Request, file: UploadFile = File(...)):
    _guard_data_ops(request)
    raw = await file.read()
    try:
        sql = raw.decode("utf-8")
    except UnicodeDecodeError:
        raise HTTPException(400, "El archivo no es texto SQL válido (UTF-8).")
    try:
        return admin.importar_sql(sql)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Error al ejecutar el SQL: {e}")


# ---- Copias de seguridad (respaldos con fecha) ----
@router.get("/admin/respaldos")
def admin_respaldos():
    return admin.listar_respaldos()


@router.post("/admin/respaldos")
def admin_crear_respaldo(request: Request):
    _guard_data_ops(request)
    return admin.crear_respaldo()


@router.post("/admin/respaldos/restaurar")
def admin_restaurar_respaldo(request: Request, payload: dict):
    _guard_data_ops(request)
    try:
        return admin.restaurar_respaldo(payload.get("archivo") or "")
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        raise HTTPException(400, f"Error al restaurar: {e}")
